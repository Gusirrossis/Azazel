"""Plugin de hojas de cálculo: XLSX (openpyxl, read_only — no carga el libro entero)."""

from __future__ import annotations

from typing import Any

from . import ContextoExtraccion, ResultadoExtraccion, registrar


@registrar("application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
def extraer_xlsx(ctx: ContextoExtraccion) -> ResultadoExtraccion:
    from openpyxl import load_workbook  # type: ignore[import-untyped]

    libro = load_workbook(ctx.fuente, read_only=True, data_only=True)
    try:
        hojas: dict[str, Any] = {}
        for ws in libro.worksheets[:50]:
            hojas[ws.title] = {
                "filas": int(ws.max_row or 0),
                "columnas": int(ws.max_column or 0),
            }
        campos = {
            "hojas": sorted(hojas.keys()),
            "hojas_total": len(libro.worksheets),
            "hojas_detalle": hojas,
        }
        return ResultadoExtraccion(campos=campos)
    finally:
        libro.close()
